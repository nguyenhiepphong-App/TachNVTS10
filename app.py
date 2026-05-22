import streamlit as st
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
import io
import re
from collections import defaultdict

st.set_page_config(page_title="Tiện ích xử lý nguyện vọng tuyển sinh vào lớp 10", layout="wide")

# Tiêu đề và giới thiệu chuẩn tác phong hành chính
st.title("🎯 Tiện ích Tách và thống kê nguyện vọng ĐK tuyển sinh vào lớp 10")
st.subheader("📋 Dành cho Hội đồng thi TS (trường THPT)")

st.markdown("""
---
* **File Input:** File Excel gốc tải ra từ hệ thống [ts10.phutho.edu.vn](https://ts10.phutho.edu.vn/)
* **File Output:**
    * **File 1:** File Excel tách các cột Nguyện vọng (NV1 đến NV6, bổ sung cột Tổng NV và kẻ viền đồng bộ bảng dữ liệu).
    * **File 2:** File Excel thống kê số lượng nguyện vọng đăng ký chi tiết theo từng đơn vị trường học (Sắp xếp giảm dần từ NV1 đến NV6).
---
""")

uploaded_file = st.file_uploader("Tải lên file Excel danh sách nguyện vọng gốc (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    # --- ĐỌC VÀ XỬ LÝ FILE 1: FILE DANH SÁCH HỌC SINH ---
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    thin_side = Side(style='thin', color="000000")
    border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    header_row_idx = None
    col_nv_idx = None
    col_ghichu_idx = None
    max_col = ws.max_column
    max_row = ws.max_row

    # Bước 1: Quét tìm hàng tiêu đề, xác định vị trí cột Nguyện vọng và cột Ghi chú
    for r in range(1, min(40, max_row + 1)):
        for c in range(1, max_col + 1):
            val = str(ws.cell(row=r, column=c).value or '').lower()
            if 'nguyện vọng' in val or 'nguyenvong' in val:
                header_row_idx = r
                col_nv_idx = c
            if 'ghi chú' in val or 'ghichu' in val:
                col_ghichu_idx = c
        if header_row_idx:
            break

    if header_row_idx and col_nv_idx:
        # CỐ ĐỊNH VỊ TRÍ TUYỆT ĐỐI: Cứ sau cột Ghi chú là chèn 7 cột mới (không phụ thuộc dữ liệu hàng)
        if col_ghichu_idx:
            start_insert_col = col_ghichu_idx + 1
        else:
            start_insert_col = col_nv_idx + 1
        
        # Xác định chính xác tổng số cột cần quét kẻ viền sau khi đã cộng thêm 7 cột mới
        total_columns_fixed = start_insert_col + 7 - 1

        # Tạo tiêu đề cố định cho 7 cột mới
        new_headers = ["NV1", "NV2", "NV3", "NV4", "NV5", "NV6", "Tổng NV"]
        for i, h_text in enumerate(new_headers):
            cell = ws.cell(row=header_row_idx, column=start_insert_col + i)
            cell.value = h_text
            cell.font = Font(bold=True, name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

        thong_ke_data = defaultdict(lambda: [0] * 6)
        count_processed = 0

        # Duyệt từng hàng dữ liệu học sinh để bóc tách nguyện vọng
        for r in range(header_row_idx + 1, max_row + 1):
            nv_text_raw = ws.cell(row=r, column=col_nv_idx).value
            nv_text = str(nv_text_raw or '').strip()
            
            lines = [line.strip() for line in nv_text.split('\n') if line.strip()]
            total_nv = len(lines)
            
            if nv_text_raw: 
                count_processed += 1

            # Ghi dữ liệu vào đúng vị trí 6 cột NV mới tính từ mốc cố định sau Ghi chú
            for index in range(6):
                target_cell = ws.cell(row=r, column=start_insert_col + index)
                
                if index < total_nv:
                    line = lines[index]
                    match = re.match(r'^[\d\.]+\s*(.*)$', line)
                    clean_nv = match.group(1).strip() if match else line
                    
                    target_cell.value = clean_nv
                    if clean_nv:
                        thong_ke_data[clean_nv][index] += 1
                else:
                    # Ghi đè ô trống để dọn dẹp sạch sẽ các cột ẩn hoặc dữ liệu rác cũ nếu có
                    target_cell.value = ""
                
                target_cell.font = Font(name="Times New Roman", size=11)
                target_cell.border = border_style
                target_cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Ghi dữ liệu vào cột Tổng NV
            total_cell = ws.cell(row=r, column=start_insert_col + 6)
            total_cell.value = total_nv if nv_text_raw else ""
            total_cell.font = Font(bold=True, name="Times New Roman", size=11)
            total_cell.border = border_style
            total_cell.alignment = Alignment(horizontal="center", vertical="center")

            # ĐỒNG BỘ: Kẻ đường viền đen khít toàn bộ bảng từ cột 1 đến hết cột Tổng NV
            for c in range(1, total_columns_fixed + 1):
                ws.cell(row=r, column=c).border = border_style

        output_ds = io.BytesIO()
        wb.save(output_ds)

        # --- KHỞI TẠO VÀ XỬ LÝ FILE 2: FILE THỐNG KÊ TỔNG HỢP ---
        wb_tk = openpyxl.Workbook()
        ws_tk = wb_tk.active
        ws_tk.title = "Thong Ke Nguyen Vong"
        
        tk_headers = ["STT", "Tên trường", "Số NV1", "Số NV2", "Số NV3", "Số NV4", "Số NV5", "Số NV6"]
        for col_num, header_title in enumerate(tk_headers, 1):
            c_cell = ws_tk.cell(row=1, column=col_num)
            c_cell.value = header_title
            c_cell.font = Font(bold=True, name="Times New Roman", size=11)
            c_cell.alignment = Alignment(horizontal="center", vertical="center")
            c_cell.border = border_style

        # Sắp xếp đa tiêu chí giảm dần từ NV1 đến NV6 (Trường nhiều NV1 lên đầu)
        sorted_truong = sorted(
            thong_ke_data.items(), 
            key=lambda x: (-x[1][0], -x[1][1], -x[1][2], -x[1][3], -x[1][4], -x[1][5])
        )

        row_tk_idx = 2
        stt_counter = 1
        
        for truong_ten, nv_counts in sorted_truong:
            if not truong_ten: continue
            
            ws_tk.cell(row=row_tk_idx, column=1, value=stt_counter).alignment = Alignment(horizontal="center")
            ws_tk.cell(row=row_tk_idx, column=2, value=truong_ten)
            
            for nv_i in range(6):
                val_count = nv_counts[nv_i]
                ws_tk.cell(row=row_tk_idx, column=3 + nv_i, value=val_count if val_count > 0 else "")
                ws_tk.cell(row=row_tk_idx, column=3 + nv_i).alignment = Alignment(horizontal="center")
            
            for c_idx in range(1, 9):
                cell_obj = ws_tk.cell(row=row_tk_idx, column=c_idx)
                cell_obj.font = Font(name="Times New Roman", size=11)
                cell_obj.border = border_style
                
            stt_counter += 1
            row_tk_idx += 1

        # Dòng tổng kết cuối bảng thống kê
        ws_tk.cell(row=row_tk_idx, column=1, value="Tổng").font = Font(bold=True, name="Times New Roman", size=11)
        ws_tk.cell(row=row_tk_idx, column=1).alignment = Alignment(horizontal="center")
        ws_tk.cell(row=row_tk_idx, column=1).border = border_style
        ws_tk.cell(row=row_tk_idx, column=2).border = border_style
        
        for nv_i in range(6):
            col_letter = openpyxl.utils.get_column_letter(3 + nv_i)
            sum_formula = f"=SUM({col_letter}2:{col_letter}{row_tk_idx-1})"
            
            sum_cell = ws_tk.cell(row=row_tk_idx, column=3 + nv_i, value=sum_formula)
            sum_cell.font = Font(bold=True, name="Times New Roman", size=11)
            sum_cell.alignment = Alignment(horizontal="center")
            sum_cell.border = border_style

        max_len_truong = max([len(t) for t in thong_ke_data.keys()] + [20])
        ws_tk.column_dimensions['B'].width = max_len_truong + 3
        ws_tk.column_dimensions['A'].width = 6

        output_tk = io.BytesIO()
        wb_tk.save(output_tk)

        st.success(f"Xử lý dữ liệu thành công. Đã hoàn thành báo cáo dữ liệu cho {count_processed} thí sinh đăng ký.")
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            st.info("📄 File 1: Excel Tách các cột Nguyện vọng")
            st.download_button(
                label="📥 Tải file Excel Danh Sách",
                data=output_ds.getvalue(),
                file_name="1_Danh_Sach_Tach_Cot_NV_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        with col_btn2:
            st.success("📊 File 2: Excel thống kê nguyện vọng từng trường")
            st.download_button(
                label="📥 Tải file Excel Thống Kê Trường",
                data=output_tk.getvalue(),
                file_name="2_Thong_Ke_Nguyen_Vong_Truong_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.error("Cấu trúc file không hợp lệ (Không tìm thấy cột dữ liệu Nguyện vọng).")

# Thêm thông tin tác giả và bản quyền ở chân trang ứng dụng theo mẫu
st.markdown("<br><br><br>", unsafe_allow_html=True)
st.markdown("""
---
<div style='text-align: left; color: #555555; font-size: 14px;'>
    Nguyen Hiep Phong – THPT Bến Tre
</div>
""", unsafe_allow_html=True)
